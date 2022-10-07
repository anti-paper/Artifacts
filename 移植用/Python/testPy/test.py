#作成日：2021/03/07（日）
#作成者：中島 健一

#更新日：2021/05/20（木）
#更新者：中島 健一
#num=256**2-1

#print("num(10)=% d"%num)
#import os,csv
'''
i=1

with open('test.txt') as t:
    texts=t.read().split()

for txt in texts:
    fileStr=str(i)+'_'+txt

    if not os.path.exists(fileStr):
        os.mkdir(fileStr)
    
    i+=1
'''
#更新日：2021/05/20（木）
#更新者：中島 健一
'''
with open('test.csv') as f:
    texts=csv.reader(f)

    for row in texts:
        folderStr=row[0]+'_'+row[1]

        if not os.path.exists(folderStr):
            os.mkdir(folderStr)
'''
#更新日：2021/05/21（金）
#更新者：中島 健一
'''
testarray=['aa','aa','ab','ba','ab','bb']
testdictionary={}

for test in testarray:
    testdictionary.setdefault(test,0)
    testdictionary[test]+=1

for key,value in testdictionary.items():
    print('{}:{}'.format(key,value))
'''
'''
import csv

with open('test.csv') as f:
    a=csv.reader(f)
    h=next(a)
    b={}

    for c in a:
        b.setdefault(c[0],0)
        b[c[0]]+=1

print(h)

for key,value in b.items():
    print('{}:{}'.format(key,value))
'''
#更新日：2021/05/22（土）
#更新者：中島 健一
'''
import pyperclip

delim='\r\n'
lines=pyperclip.paste().split(delim)
text=''

for i,line in enumerate(lines):
    #text=text+str(i+1)+'> '+line+delim
    text='{}{}> {}{}'.format(text,i+1,line,delim)

pyperclip.copy(text)
'''
'''
import pyperclip

lines=pyperclip.paste().split('\r\n')
text=''

for i in range(0,len(lines),2):
    text='{}{}\t{}\n'.format(text,lines[i],lines[i+1])

with open('methods.txt','a') as f:
    f.write(text)
'''
'''
import csv,pyperclip

lines=pyperclip.paste().split('\r\n')
text=''

with open('methods.csv','a',newline='') as f:
    writer=csv.writer(f)

    for i in range(0,len(lines),2):
        writer.writerow([lines[i],lines[i+1]])
'''
'''
import webbrowser

webbrowser.open('https://tonari-it.com/python-webbrowser-open/')
'''
'''
import openpyxl

wb=openpyxl.load_workbook('test_sample.xlsx')
sheet=wb.worksheets[0]

for cell in list(sheet.columns)[1]:
    print(cell.value)
'''
'''
import openpyxl,glob

wb=openpyxl.Workbook()
sheet=wb.active
sheet.title='exceler'
wb.save('test.xlsx')
print(glob.glob('*.xlsx'))
'''
'''
import openpyxl

fName='test.xlsx'
wb=openpyxl.load_workbook(fName)
ws=wb.active

for i in range(10):
    ws.append([i,i+1])

wb.save(fName)
'''
#更新日：2021/05/23（日）
#更新者：中島 健一
'''
import zipfile,os,datetime
'''
'''
for folder,subfolders,files in os.walk('test'):
    print('folder:{}'.format(folder))
    print('subfolders:{}'.format(subfolders))
    print('files:{}'.format(files))
    print()
'''
'''
now=datetime.datetime.now()
nameStr='test_{0:%Y%m%d}.zip'.format(now)

with zipfile.ZipFile(nameStr,'w') as myzip:
    for folder,subfolders,files in os.walk('test'):
        myzip.write(folder)

        for file in files:
            myzip.write(os.path.join(folder,file))
'''
'''
print(os.path.getctime('test'))
'''
'''
from operator import itemgetter, truediv

files=os.listdir()
xlsxs=[]

for file in files:
    base,ext=os.path.splitext(file)

    if ext=='.xlsx':
        xlsxs.append([file,os.path.getctime(file)])

xlsxs.sort(key=itemgetter(1),reverse=True)

maxCnt=1

for i,xlsx in enumerate(xlsxs):
    if i>maxCnt-1:
        print('{} is going to deleted'.format(xlsx[0]))
'''
'''
import requests

res=requests.get('https://tonari-it.com/python-requests-get-response/')
#print(res.text)
with open('reshtml.html','w',encoding='utf-8') as f:
    f.write(res.text)
'''

#更新日：2021/05/25（火）
#更新者：中島 健一
import requests,bs4

res=requests.get('https://tonari-it.com')
res.raise_for_status()
soup=bs4.BeautifulSoup(res.text,"html.parser")
elems=soup.select('#list a')

with open(r'C:\Users\nakaj\OneDrive\まとめ\勉強\プログラミング\TEST\Python\testPy\test.txt','w') as f:
    for elem in elems:
        f.write('{}({})\n'.format(elem.get('title'),elem.get('href')))