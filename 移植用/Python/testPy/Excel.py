import openpyxl

link='test.xlsx'
wb=openpyxl.load_workbook(link)
ws=wb.active
for index,cell in enumerate(list(ws.columns)[0]):
    if index==0:
        print([ws.cell(index+1,1).value,ws.cell(index+1,2).value])
    else:
        print(ws.cell(index+1,1).value,end=':')
        print(ws.cell(index+1,2).value)